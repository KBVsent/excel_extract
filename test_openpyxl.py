#!/usr/bin/env python3
"""
Excel Content Extractor using openpyxl
Extracts all content from Excel files including:
- Cell values and formulas
- Cell formatting (font, fill, alignment, borders)
- Images
- Charts
- Comments
- Merged cells
- Hyperlinks
"""

import os
import json
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string


class ExcelExtractor:
    """Extract all content from Excel files"""
    
    def __init__(self, excel_path: str, output_dir: str = "extracted_content"):
        """
        Initialize the extractor
        
        Args:
            excel_path: Path to the Excel file
            output_dir: Directory to save extracted content
        """
        self.excel_path = excel_path
        self.output_dir = output_dir
        self.workbook = None
        self.extracted_data = {}
        
        # Create output directory
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)
        
    def load_excel(self):
        """Load the Excel workbook"""
        print(f"Loading Excel file: {self.excel_path}")
        self.workbook = load_workbook(
            self.excel_path,
            data_only=False,  # Keep formulas
            keep_vba=True,    # Keep VBA macros if present
        )
        print(f"Workbook loaded successfully with {len(self.workbook.sheetnames)} sheets")
        
    def extract_cell_value(self, cell) -> Dict[str, Any]:
        """Extract cell value and metadata"""
        cell_data = {
            "value": cell.value,
            "data_type": cell.data_type,
            "number_format": cell.number_format,
        }
        
        # Extract formula if present
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            cell_data["formula"] = cell.value
            
        return cell_data
    
    def extract_cell_style(self, cell) -> Dict[str, Any]:
        """Extract cell styling information"""
        style_data = {}
        
        # Font information
        if cell.font:
            style_data["font"] = {
                "name": cell.font.name,
                "size": cell.font.size,
                "bold": cell.font.bold,
                "italic": cell.font.italic,
                "underline": cell.font.underline,
                "color": str(cell.font.color.rgb) if cell.font.color else None,
            }
        
        # Fill/Background color
        if cell.fill:
            style_data["fill"] = {
                "fill_type": cell.fill.fill_type,
                "start_color": str(cell.fill.start_color.rgb) if cell.fill.start_color else None,
                "end_color": str(cell.fill.end_color.rgb) if cell.fill.end_color else None,
            }
        
        # Alignment
        if cell.alignment:
            style_data["alignment"] = {
                "horizontal": cell.alignment.horizontal,
                "vertical": cell.alignment.vertical,
                "wrap_text": cell.alignment.wrap_text,
                "text_rotation": cell.alignment.text_rotation,
            }
        
        # Border
        if cell.border:
            style_data["border"] = {
                "left": str(cell.border.left.style) if cell.border.left else None,
                "right": str(cell.border.right.style) if cell.border.right else None,
                "top": str(cell.border.top.style) if cell.border.top else None,
                "bottom": str(cell.border.bottom.style) if cell.border.bottom else None,
            }
        
        return style_data
    
    def extract_merged_cells(self, worksheet: Worksheet) -> List[str]:
        """Extract merged cell ranges"""
        merged_cells = []
        for merged_range in worksheet.merged_cells.ranges:
            merged_cells.append(str(merged_range))
        return merged_cells
    
    def extract_hyperlinks(self, worksheet: Worksheet) -> Dict[str, str]:
        """Extract hyperlinks from cells"""
        hyperlinks = {}
        for cell in worksheet._cells.values():
            if cell.hyperlink:
                hyperlinks[cell.coordinate] = {
                    "url": cell.hyperlink.target,
                    "display": cell.value,
                    "tooltip": cell.hyperlink.tooltip if hasattr(cell.hyperlink, 'tooltip') else None,
                }
        return hyperlinks
    
    def extract_comments(self, worksheet: Worksheet) -> Dict[str, Dict[str, str]]:
        """Extract comments from cells"""
        comments = {}
        for cell in worksheet._cells.values():
            if cell.comment:
                comments[cell.coordinate] = {
                    "text": cell.comment.text,
                    "author": cell.comment.author,
                }
        return comments
    
    def _parse_drawing_positions(self, sheet_name: str) -> Dict[str, Tuple[int, int, int, int]]:
        """Parse drawing.xml to get image positions
        Returns: dict mapping image file to (from_col, from_row, to_col, to_row)
        """
        positions = {}
        
        try:
            # Find the sheet's relationship ID
            with zipfile.ZipFile(self.excel_path, 'r') as zip_ref:
                # Read workbook.xml.rels to find sheet file
                # Then read sheet's .rels to find drawing file
                # Then parse drawing.xml
                
                # For simplicity, try common drawing files
                for drawing_idx in range(1, 10):
                    drawing_file = f'xl/drawings/drawing{drawing_idx}.xml'
                    if drawing_file not in zip_ref.namelist():
                        continue
                    
                    # Parse the drawing XML
                    content = zip_ref.read(drawing_file)
                    root = ET.fromstring(content)
                    
                    # Define namespace
                    ns = {'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'}
                    
                    # Find all anchor elements with images
                    img_idx = 1
                    for anchor in root.findall('.//xdr:twoCellAnchor', ns) + root.findall('.//xdr:oneCellAnchor', ns):
                        # Get from position
                        from_elem = anchor.find('.//xdr:from', ns)
                        if from_elem is not None:
                            from_col = int(from_elem.find('xdr:col', ns).text)
                            from_row = int(from_elem.find('xdr:row', ns).text)
                            
                            # Get to position (if two-cell anchor)
                            to_col, to_row = from_col, from_row
                            to_elem = anchor.find('.//xdr:to', ns)
                            if to_elem is not None:
                                to_col = int(to_elem.find('xdr:col', ns).text)
                                to_row = int(to_elem.find('xdr:row', ns).text)
                            
                            # Store position (use generic key since we don't have image-to-position mapping yet)
                            positions[f'image_{img_idx}'] = (from_col, from_row, to_col, to_row)
                            img_idx += 1
        except Exception as e:
            print(f"  Warning: Could not parse drawing positions: {e}")
        
        return positions
    
    def extract_images(self, worksheet: Worksheet, sheet_name: str) -> List[Dict[str, Any]]:
        """Extract images from worksheet"""
        images_info = []
        
        # Get image positions from drawing.xml
        image_positions = self._parse_drawing_positions(sheet_name)
        
        # Method 1: Try standard _images attribute
        if hasattr(worksheet, '_images') and worksheet._images:
            for idx, image in enumerate(worksheet._images):
                # Save image to file
                image_filename = f"{sheet_name}_image_{idx + 1}.{image.format}"
                image_path = os.path.join(self.output_dir, image_filename)
                
                with open(image_path, "wb") as img_file:
                    img_file.write(image._data())
                
                # Get position if available
                pos_key = f'image_{idx + 1}'
                position = image_positions.get(pos_key)
                
                images_info.append({
                    "filename": image_filename,
                    "format": image.format,
                    "anchor": str(image.anchor) if hasattr(image, 'anchor') else None,
                    "width": image.width if hasattr(image, 'width') else None,
                    "height": image.height if hasattr(image, 'height') else None,
                    "position": position,  # (from_col, from_row, to_col, to_row)
                })
                print(f"  Extracted image: {image_filename}")
        
        # Method 2: Extract directly from ZIP file (for all sheets on first call)
        # Only do this once for the workbook
        if not images_info and not hasattr(self, '_images_extracted_from_zip'):
            self._images_extracted_from_zip = True
            try:
                with zipfile.ZipFile(self.excel_path, 'r') as zip_ref:
                    # Find all media files
                    media_files = [f for f in zip_ref.namelist() if 'xl/media/' in f and not f.endswith('/')]
                    
                    if media_files:
                        print(f"  Found {len(media_files)} image(s) in workbook")
                        
                        for idx, media_file in enumerate(media_files, 1):
                            # Extract file extension
                            file_ext = os.path.splitext(media_file)[1][1:]  # Remove the dot
                            if not file_ext:
                                file_ext = 'png'  # default
                            
                            # Use generic naming since we can't determine which sheet
                            image_filename = f"workbook_image_{idx}.{file_ext}"
                            image_path = os.path.join(self.output_dir, image_filename)
                            
                            # Extract and save image
                            with open(image_path, 'wb') as img_file:
                                img_file.write(zip_ref.read(media_file))
                            
                            # Get position if available
                            pos_key = f'image_{idx}'
                            position = image_positions.get(pos_key)
                            
                            # Get image info
                            images_info.append({
                                "filename": image_filename,
                                "format": file_ext.upper(),
                                "source": media_file,
                                "anchor": None,
                                "width": None,
                                "height": None,
                                "position": position,  # (from_col, from_row, to_col, to_row)
                            })
                            print(f"  Extracted image: {image_filename} at position {position}")
            except Exception as e:
                print(f"  Warning: Could not extract images from ZIP: {e}")
        
        return images_info
        
        return images_info
    
    def extract_charts(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        """Extract chart information"""
        charts_info = []
        
        if hasattr(worksheet, '_charts') and worksheet._charts:
            for idx, chart in enumerate(worksheet._charts):
                # Extract chart title - it's often a complex nested structure
                chart_title = None
                if hasattr(chart, 'title') and chart.title:
                    # Try to extract title text from nested structure
                    try:
                        if hasattr(chart.title, 'tx') and chart.title.tx:
                            if hasattr(chart.title.tx, 'rich') and chart.title.tx.rich:
                                # Extract text from rich text paragraphs
                                if hasattr(chart.title.tx.rich, 'p'):
                                    paragraphs = chart.title.tx.rich.p
                                    texts = []
                                    for p in paragraphs:
                                        if hasattr(p, 'r'):  # text runs
                                            for run in p.r:
                                                if hasattr(run, 't'):
                                                    texts.append(run.t)
                                    if texts:
                                        chart_title = ''.join(texts)
                    except:
                        chart_title = None
                
                chart_data = {
                    "type": chart.__class__.__name__,
                    "title": chart_title,
                    "anchor": None,  # Simplified - anchor is complex object
                }
                charts_info.append(chart_data)
                print(f"  Found chart: {chart_data['type']}" + (f" - '{chart_title}'" if chart_title else ""))
        
        return charts_info
    
    def extract_sheet_properties(self, worksheet: Worksheet) -> Dict[str, Any]:
        """Extract worksheet properties"""
        properties = {
            "title": worksheet.title,
            "max_row": worksheet.max_row,
            "max_column": worksheet.max_column,
            "sheet_state": worksheet.sheet_state,
            "sheet_view": {
                "show_gridlines": worksheet.sheet_view.showGridLines if worksheet.sheet_view else None,
                "show_row_col_headers": worksheet.sheet_view.showRowColHeaders if worksheet.sheet_view else None,
                "zoom_scale": worksheet.sheet_view.zoomScale if worksheet.sheet_view else None,
            },
        }
        
        # Column dimensions
        if worksheet.column_dimensions:
            properties["column_widths"] = {
                col: dim.width for col, dim in worksheet.column_dimensions.items() if dim.width
            }
        
        # Row dimensions
        if worksheet.row_dimensions:
            properties["row_heights"] = {
                row: dim.height for row, dim in worksheet.row_dimensions.items() if dim.height
            }
        
        return properties
    
    def extract_worksheet(self, worksheet: Worksheet) -> Dict[str, Any]:
        """Extract all content from a worksheet"""
        sheet_name = worksheet.title
        print(f"\nExtracting sheet: {sheet_name}")
        
        sheet_data = {
            "properties": self.extract_sheet_properties(worksheet),
            "cells": {},
            "merged_cells": self.extract_merged_cells(worksheet),
            "hyperlinks": self.extract_hyperlinks(worksheet),
            "comments": self.extract_comments(worksheet),
            "images": self.extract_images(worksheet, sheet_name),
            "charts": self.extract_charts(worksheet),
        }
        
        # Extract cell data and styles
        print(f"  Extracting cells...")
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None or cell.has_style:
                    cell_key = cell.coordinate
                    sheet_data["cells"][cell_key] = {
                        "data": self.extract_cell_value(cell),
                        "style": self.extract_cell_style(cell),
                    }
        
        print(f"  Extracted {len(sheet_data['cells'])} cells")
        print(f"  Found {len(sheet_data['merged_cells'])} merged cell ranges")
        print(f"  Found {len(sheet_data['hyperlinks'])} hyperlinks")
        print(f"  Found {len(sheet_data['comments'])} comments")
        
        return sheet_data
    
    def extract_workbook_properties(self) -> Dict[str, Any]:
        """Extract workbook-level properties"""
        properties = {
            "sheetnames": self.workbook.sheetnames,
            "active_sheet": self.workbook.active.title,
        }
        
        # Document properties
        if self.workbook.properties:
            props = self.workbook.properties
            properties["document_properties"] = {
                "title": props.title,
                "subject": props.subject,
                "creator": props.creator,
                "keywords": props.keywords,
                "description": props.description,
                "last_modified_by": props.lastModifiedBy,
                "created": str(props.created) if props.created else None,
                "modified": str(props.modified) if props.modified else None,
            }
        
        # Named ranges
        if self.workbook.defined_names:
            properties["named_ranges"] = {
                name: str(defn.attr_text) for name, defn in self.workbook.defined_names.items()
            }
        
        return properties
    
    def extract_all(self) -> Dict[str, Any]:
        """Extract all content from the workbook"""
        self.load_excel()
        
        print("\n" + "="*60)
        print("Starting full extraction")
        print("="*60)
        
        self.extracted_data = {
            "workbook_properties": self.extract_workbook_properties(),
            "sheets": {}
        }
        
        # Extract each worksheet
        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]
            self.extracted_data["sheets"][sheet_name] = self.extract_worksheet(worksheet)
        
        print("\n" + "="*60)
        print("Extraction complete!")
        print("="*60)
        
        return self.extracted_data
    
    def extract_single_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """Extract content from a single sheet
        
        Args:
            sheet_name: Name of the sheet to extract
            
        Returns:
            Dictionary containing sheet data
        """
        self.load_excel()
        
        # Check if sheet exists
        if sheet_name not in self.workbook.sheetnames:
            available_sheets = ", ".join(self.workbook.sheetnames)
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {available_sheets}")
        
        print("\n" + "="*60)
        print(f"Extracting single sheet: {sheet_name}")
        print("="*60)
        
        # Extract only the specified sheet
        worksheet = self.workbook[sheet_name]
        sheet_data = self.extract_worksheet(worksheet)
        
        # Create minimal workbook properties
        self.extracted_data = {
            "workbook_properties": {
                "sheetnames": [sheet_name],
                "active_sheet": sheet_name,
                "source_file": self.excel_path,
            },
            "sheets": {
                sheet_name: sheet_data
            }
        }
        
        print("\n" + "="*60)
        print("Extraction complete!")
        print("="*60)
        
        return self.extracted_data
    
    def list_sheets(self) -> List[str]:
        """List all sheet names in the workbook
        
        Returns:
            List of sheet names
        """
        if not self.workbook:
            self.load_excel()
        return self.workbook.sheetnames
    
    def save_to_json(self, output_filename: str = "extracted_data.json"):
        """Save extracted data to JSON file"""
        output_path = os.path.join(self.output_dir, output_filename)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.extracted_data, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"\nExtracted data saved to: {output_path}")
        return output_path
    
    def generate_summary(self) -> str:
        """Generate a human-readable summary of extracted content"""
        summary = []
        summary.append("\n" + "="*60)
        summary.append("EXTRACTION SUMMARY")
        summary.append("="*60)
        
        # Workbook info
        wb_props = self.extracted_data.get("workbook_properties", {})
        summary.append(f"\nWorkbook: {self.excel_path}")
        summary.append(f"Active Sheet: {wb_props.get('active_sheet', 'N/A')}")
        summary.append(f"Total Sheets: {len(self.extracted_data.get('sheets', {}))}")
        
        # Sheet details
        for sheet_name, sheet_data in self.extracted_data.get("sheets", {}).items():
            summary.append(f"\n--- Sheet: {sheet_name} ---")
            props = sheet_data.get("properties", {})
            summary.append(f"  Dimensions: {props.get('max_row', 0)} rows × {props.get('max_column', 0)} columns")
            summary.append(f"  Cells with data: {len(sheet_data.get('cells', {}))}")
            summary.append(f"  Merged cells: {len(sheet_data.get('merged_cells', []))}")
            summary.append(f"  Hyperlinks: {len(sheet_data.get('hyperlinks', {}))}")
            summary.append(f"  Comments: {len(sheet_data.get('comments', {}))}")
            summary.append(f"  Images: {len(sheet_data.get('images', []))}")
            summary.append(f"  Charts: {len(sheet_data.get('charts', []))}")
        
        summary.append("\n" + "="*60)
        
        summary_text = "\n".join(summary)
        print(summary_text)
        
        # Save summary to file
        summary_path = os.path.join(self.output_dir, "extraction_summary.txt")
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write(summary_text)
        
        return summary_text
    
    def generate_markdown(self, output_filename: str = "extracted_content.md") -> str:
        """Generate a human-readable Markdown report of extracted content"""
        md_lines = []
        
        # Title
        md_lines.append(f"# {os.path.basename(self.excel_path)}")
        md_lines.append("")
        
        # Process each sheet
        for sheet_name, sheet_data in self.extracted_data.get("sheets", {}).items():
            md_lines.append(f"## {sheet_name}")
            md_lines.append("")
            
            # Get images for this sheet
            images = sheet_data.get("images", [])
            image_by_row = {}  # Map row number -> list of images starting at that row
            
            for img in images:
                position = img.get("position")
                if position:
                    from_col, from_row, to_col, to_row = position
                    # Store images by their starting row (Excel uses 0-indexed internally)
                    # Add 1 to convert to 1-indexed for display
                    display_row = from_row + 1
                    if display_row not in image_by_row:
                        image_by_row[display_row] = []
                    
                    col_letter_from = get_column_letter(from_col + 1)
                    col_letter_to = get_column_letter(to_col + 1)
                    row_to = to_row + 1
                    
                    image_by_row[display_row].append({
                        'filename': img['filename'],
                        'position_text': f"{col_letter_from}{display_row} to {col_letter_to}{row_to}"
                    })
            
            # Cell data as table
            cells = sheet_data.get("cells", {})
            if cells:
                # Organize cells by row
                rows_dict = {}
                for cell_ref, cell_data in cells.items():
                    # Parse cell reference (e.g., "A1" -> row=1, col=A)
                    col = ''.join(filter(str.isalpha, cell_ref))
                    row = int(''.join(filter(str.isdigit, cell_ref)))
                    
                    if row not in rows_dict:
                        rows_dict[row] = {}
                    rows_dict[row][col] = cell_data
                
                # Find all unique columns
                all_cols = set()
                for row_data in rows_dict.values():
                    all_cols.update(row_data.keys())
                all_cols = sorted(all_cols)
                
                # Generate table
                sorted_rows = sorted(rows_dict.keys())
                
                if sorted_rows and all_cols:
                    # Table header
                    md_lines.append("| " + " | ".join(all_cols) + " |")
                    md_lines.append("|" + "|".join(["---"] * len(all_cols)) + "|")
                    
                    # Table rows - show images after their corresponding rows
                    last_row = 0
                    for row_num in sorted_rows:
                        # First, show any images that appear between last_row and current row
                        for check_row in range(last_row + 1, row_num):
                            if check_row in image_by_row:
                                md_lines.append("")  # End table temporarily
                                for img_info in image_by_row[check_row]:
                                    md_lines.append(f"**Image at {img_info['position_text']}:**")
                                    md_lines.append("")
                                    md_lines.append(f"![Image]({img_info['filename']})")
                                    md_lines.append("")
                                # Restart table
                                md_lines.append("| " + " | ".join(all_cols) + " |")
                                md_lines.append("|" + "|".join(["---"] * len(all_cols)) + "|")
                        
                        # Show the data row
                        row_data = rows_dict[row_num]
                        values = []
                        for col in all_cols:
                            if col in row_data:
                                value = row_data[col]['data']['value']
                                # Format value for markdown
                                if value is None:
                                    value = ""
                                elif isinstance(value, str):
                                    # Escape pipes
                                    value = value.replace("|", "\\|").replace("\n", " ")
                                else:
                                    # Convert dates and other types to string
                                    value = str(value)
                                    # Clean up datetime string
                                    if "datetime" in value:
                                        value = value.replace("datetime.datetime(", "").replace(")", "")
                                values.append(value)
                            else:
                                values.append("")
                        md_lines.append("| " + " | ".join(values) + " |")
                        
                        # Check if there's an image starting at the next row
                        if row_num + 1 in image_by_row:
                            md_lines.append("")  # End table
                            for img_info in image_by_row[row_num + 1]:
                                md_lines.append(f"**Image at {img_info['position_text']}:**")
                                md_lines.append("")
                                md_lines.append(f"![Image]({img_info['filename']})")
                                md_lines.append("")
                            # Continue table if there are more rows
                            if row_num < sorted_rows[-1]:
                                md_lines.append("| " + " | ".join(all_cols) + " |")
                                md_lines.append("|" + "|".join(["---"] * len(all_cols)) + "|")
                        
                        last_row = row_num
                    
                    # Check for images after the last data row
                    max_row = sorted_rows[-1] if sorted_rows else 0
                    for check_row in range(max_row + 1, max(image_by_row.keys()) + 1 if image_by_row else max_row + 1):
                        if check_row in image_by_row:
                            md_lines.append("")
                            for img_info in image_by_row[check_row]:
                                md_lines.append(f"**Image at {img_info['position_text']}:**")
                                md_lines.append("")
                                md_lines.append(f"![Image]({img_info['filename']})")
                                md_lines.append("")
                    
                    md_lines.append("")
            elif images:
                # No cell data, but there are images
                for img in images:
                    position = img.get("position")
                    if position:
                        from_col, from_row, to_col, to_row = position
                        col_letter_from = get_column_letter(from_col + 1)
                        col_letter_to = get_column_letter(to_col + 1)
                        row_from = from_row + 1
                        row_to = to_row + 1
                        md_lines.append(f"**Image at {col_letter_from}{row_from} to {col_letter_to}{row_to}:**")
                    md_lines.append("")
                    md_lines.append(f"![Image]({img['filename']})")
                    md_lines.append("")
            
            # Hyperlinks (if any)
            hyperlinks = sheet_data.get("hyperlinks", {})
            if hyperlinks:
                md_lines.append("### Hyperlinks")
                md_lines.append("")
                for cell_ref, link_data in hyperlinks.items():
                    display = link_data.get('display', '')
                    url = link_data.get('url', '')
                    md_lines.append(f"- **{cell_ref}**: [{display}]({url})")
                md_lines.append("")
            
            # Comments (if any)
            comments = sheet_data.get("comments", {})
            if comments:
                md_lines.append("### Comments")
                md_lines.append("")
                for cell_ref, comment_data in comments.items():
                    author = comment_data.get('author', '')
                    text = comment_data.get('text', '').replace("\n", " ")
                    md_lines.append(f"- **{cell_ref}** ({author}): {text}")
                md_lines.append("")
        
        markdown_content = "\n".join(md_lines)
        
        # Save to file
        output_path = os.path.join(self.output_dir, output_filename)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(markdown_content)
        
        print(f"\nMarkdown report saved to: {output_path}")
        return markdown_content


def main():
    """Main function to run the extractor"""
    import sys
    import argparse
    
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Extract content from Excel files')
    parser.add_argument('excel_file', nargs='?', help='Path to Excel file')
    parser.add_argument('--sheet', '-s', help='Extract only the specified sheet')
    parser.add_argument('--list-sheets', '-l', action='store_true', help='List all sheets and exit')
    parser.add_argument('--output-dir', '-o', help='Output directory (default: extracted_content)')
    
    args = parser.parse_args()
    
    # Determine Excel file
    excel_file = args.excel_file
    if not excel_file:
        # Default: look for any Excel file in current directory
        excel_files = list(Path('.').glob('*.xlsx')) + list(Path('.').glob('*.xlsm'))
        if excel_files:
            excel_file = str(excel_files[0])
            print(f"No file specified, using: {excel_file}")
        else:
            print("Usage: python test_openpyxl.py <excel_file.xlsx> [--sheet SHEET_NAME]")
            print("\nOptions:")
            print("  --sheet, -s SHEET_NAME    Extract only the specified sheet")
            print("  --list-sheets, -l         List all sheets in the workbook")
            print("  --output-dir, -o DIR      Specify output directory")
            print("\nExamples:")
            print("  python test_openpyxl.py data.xlsx")
            print("  python test_openpyxl.py data.xlsx --sheet Sheet1")
            print("  python test_openpyxl.py data.xlsx --list-sheets")
            return
    
    if not os.path.exists(excel_file):
        print(f"Error: File not found: {excel_file}")
        return
    
    # Determine output directory
    output_dir = args.output_dir or "extracted_content"
    
    # Create extractor
    extractor = ExcelExtractor(excel_file, output_dir=output_dir)
    
    try:
        # List sheets mode
        if args.list_sheets:
            print(f"\nSheets in {excel_file}:")
            sheets = extractor.list_sheets()
            for idx, sheet in enumerate(sheets, 1):
                print(f"  {idx}. {sheet}")
            print(f"\nTotal: {len(sheets)} sheet(s)")
            return
        
        # Extract content
        if args.sheet:
            # Extract single sheet
            try:
                extractor.extract_single_sheet(args.sheet)
                output_suffix = f"_{args.sheet}"
                
                # Save to JSON
                json_filename = f"extracted_data{output_suffix}.json"
                extractor.save_to_json(json_filename)
                
                # Generate Markdown
                md_filename = f"extracted_content{output_suffix}.md"
                extractor.generate_markdown(md_filename)
                
                # Generate summary
                extractor.generate_summary()
                
                print(f"\nExtracted content saved to: {extractor.output_dir}/")
                
            except ValueError as e:
                print(f"\nError: {e}")
                print("\nUse --list-sheets to see available sheets")
                return
        else:
            # Extract all sheets - each to a separate file
            print("\n" + "="*60)
            print("Extracting all sheets (each to separate files)")
            print("="*60)
            
            sheets = extractor.list_sheets()
            total_sheets = len(sheets)
            
            for idx, sheet_name in enumerate(sheets, 1):
                print(f"\n[{idx}/{total_sheets}] Processing sheet: {sheet_name}")
                
                try:
                    # Extract this sheet
                    extractor.extract_single_sheet(sheet_name)
                    
                    # Save to JSON
                    json_filename = f"extracted_data_{sheet_name}.json"
                    extractor.save_to_json(json_filename)
                    
                    # Generate Markdown
                    md_filename = f"extracted_content_{sheet_name}.md"
                    extractor.generate_markdown(md_filename)
                    
                    print(f"✓ Completed: {sheet_name}")
                    
                except Exception as e:
                    print(f"✗ Error processing sheet '{sheet_name}': {e}")
                    continue
            
            # Generate overall summary
            print("\n" + "="*60)
            print("ALL SHEETS EXTRACTION SUMMARY")
            print("="*60)
            print(f"Total sheets processed: {total_sheets}")
            print(f"Output directory: {extractor.output_dir}/")
            print("="*60)
        
    except Exception as e:
        print(f"Error during extraction: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
